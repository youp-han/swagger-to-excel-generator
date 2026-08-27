"""Swagger 2.0 JSON을 API 정의서 형식의 엑셀 파일로 변환한다."""

import argparse
import json
import re
import urllib.request
from copy import deepcopy

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

COLS = ["파라미터명", "파라미터 정의", "Data Type", "Data Size", "필수여부", "설명"]
COL_WIDTHS = [26, 26, 12, 10, 8, 40]

SPACER = 1  # 모든 탭의 A열은 여백으로 비워두고 B열부터 내용 시작


def C(i: int) -> int:
    """표 안에서의 1-based 열 번호를 실제 워크시트 열 번호로 변환 (A열 여백 반영)."""
    return i + SPACER


FILL_LABEL = PatternFill("solid", fgColor="FFFF00")
FILL_REQUEST = PatternFill("solid", fgColor="F2B6B0")
FILL_RESPONSE = PatternFill("solid", fgColor="B8CCE4")
FILL_SUBGROUP = PatternFill("solid", fgColor="FBE0D5")
FILL_HEADER_ROW = PatternFill("solid", fgColor="D9D9D9")
THIN = Side(style="thin", color="999999")
THICK = Side(style="thick", color="000000")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BOLD = Font(bold=True)


def apply_outer_border(ws: Worksheet, min_row: int, max_row: int, min_col: int, max_col: int):
    """지정한 범위(표 하나)의 가장자리만 두꺼운 테두리로 덮어씌운다."""
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            cell = ws.cell(row=r, column=c)
            existing = cell.border
            cell.border = Border(
                left=THICK if c == min_col else existing.left,
                right=THICK if c == max_col else existing.right,
                top=THICK if r == min_row else existing.top,
                bottom=THICK if r == max_row else existing.bottom,
            )


def fetch_swagger(url: str) -> dict:
    with urllib.request.urlopen(url, timeout=15) as resp:
        return json.loads(resp.read().decode("utf-8"))


def resolve_ref(ref: str, definitions: dict) -> dict:
    name = ref.split("/")[-1]
    return definitions.get(name, {})


def type_label(schema: dict, definitions: dict) -> str:
    if "$ref" in schema:
        return "Data"
    t = schema.get("type")
    fmt = schema.get("format")
    if t == "integer":
        return "int" if fmt in ("int32", "int64", None) else fmt
    if t == "number":
        return "decimal"
    if t == "boolean":
        return "boolean"
    if t == "string":
        if fmt == "date-time":
            return "datetime"
        return "string"
    if t == "array":
        items = schema.get("items", {})
        return f"{type_label(items, definitions)} [ ]"
    if t == "object" or not t:
        return "Data"
    return t


def flatten_schema(schema: dict, definitions: dict, level: int = 0, visited=None):
    """(level, name, definition, type, size, required, description) 튜플 목록으로 평탄화."""
    if visited is None:
        visited = set()
    rows = []

    if "$ref" in schema:
        ref_name = schema["$ref"].split("/")[-1]
        if ref_name in visited or level > 4:
            return rows
        visited = visited | {ref_name}
        resolved = resolve_ref(schema["$ref"], definitions)
        return flatten_schema(resolved, definitions, level, visited)

    if schema.get("type") == "array":
        items = schema.get("items", {})
        return flatten_schema(items, definitions, level, visited)

    props = schema.get("properties")
    if not props:
        return rows

    required = set(schema.get("required", []))
    for field_name, field_schema in props.items():
        is_ref_or_object = "$ref" in field_schema or field_schema.get("type") in ("object", "array")
        size = field_schema.get("maxLength", "")
        desc = field_schema.get("description", "")
        if field_schema.get("enum"):
            desc = (desc + " " if desc else "") + f"(허용값: {', '.join(map(str, field_schema['enum']))})"
        rows.append(
            (
                level,
                field_name,
                field_schema.get("title", ""),
                type_label(field_schema, definitions),
                size,
                "Y" if field_name in required else "N",
                desc,
            )
        )
        if is_ref_or_object:
            rows.extend(flatten_schema(field_schema, definitions, level + 1, visited))

    return rows


def indent_name(name: str, level: int) -> str:
    if level == 0:
        return name
    return ("　" * level) + "└ " + name


def sanitize_sheet_name(raw: str, used: set) -> str:
    name = re.sub(r"[\\/*?:\[\]]", "", raw)[:31]
    base = name
    i = 2
    while name in used:
        suffix = f"_{i}"
        name = base[: 31 - len(suffix)] + suffix
        i += 1
    used.add(name)
    return name


def style_label_row(ws: Worksheet, row: int, label: str, value: str, ncols: int):
    ws.cell(row=row, column=C(1), value=label).font = BOLD
    ws.cell(row=row, column=C(1)).fill = FILL_LABEL
    ws.merge_cells(start_row=row, start_column=C(2), end_row=row, end_column=C(ncols))
    vcell = ws.cell(row=row, column=C(2), value=value)
    vcell.alignment = Alignment(wrap_text=True, vertical="center")
    for c in range(1, ncols + 1):
        ws.cell(row=row, column=C(c)).border = BORDER


def style_section_banner(ws: Worksheet, row: int, title: str, fill: PatternFill, ncols: int):
    ws.merge_cells(start_row=row, start_column=C(1), end_row=row, end_column=C(ncols))
    cell = ws.cell(row=row, column=C(1), value=title)
    cell.font = BOLD
    cell.alignment = Alignment(horizontal="center")
    for c in range(1, ncols + 1):
        ws.cell(row=row, column=C(c)).fill = fill
        ws.cell(row=row, column=C(c)).border = BORDER


def style_column_header(ws: Worksheet, row: int):
    for c, name in enumerate(COLS, start=1):
        cell = ws.cell(row=row, column=C(c), value=name)
        cell.font = BOLD
        cell.fill = FILL_HEADER_ROW
        cell.alignment = Alignment(horizontal="center")
        cell.border = BORDER


def style_subgroup_row(ws: Worksheet, row: int, label: str, ncols: int):
    cell = ws.cell(row=row, column=C(1), value=label)
    cell.font = BOLD
    for c in range(1, ncols + 1):
        ws.cell(row=row, column=C(c)).fill = FILL_SUBGROUP
        ws.cell(row=row, column=C(c)).border = BORDER


def write_field_row(ws: Worksheet, row: int, level, name, definition, dtype, size, required, desc):
    values = [indent_name(name, level), definition, dtype, size, required, desc]
    for c, v in enumerate(values, start=1):
        cell = ws.cell(row=row, column=C(c), value=v)
        cell.border = BORDER
        cell.alignment = Alignment(wrap_text=True, vertical="top")


def build_detail_sheet(wb: Workbook, sheet_name: str, tag: str, path: str, method: str, op: dict, definitions: dict):
    ws = wb.create_sheet(sheet_name)
    ncols = len(COLS)
    ws.column_dimensions["A"].width = 3
    ws.row_dimensions[1].height = 6
    for i, w in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(C(i))].width = w

    title = op.get("summary") or path.rstrip("/").split("/")[-1]
    description = op.get("description", "")

    row = 2
    label_start = row
    style_label_row(ws, row, "IF Title", title, ncols); row += 1
    style_label_row(ws, row, "IF 방식 / Data 형식", "REST", ncols); row += 1
    style_label_row(ws, row, "URL", "{{HOST}}" + path, ncols); row += 1
    style_label_row(ws, row, "Method", method.upper(), ncols); row += 1
    style_label_row(ws, row, "설명", description, ncols); row += 1
    apply_outer_border(ws, label_start, row - 1, C(1), C(ncols))
    row += 1  # blank line

    # REQUEST
    request_start = row
    style_section_banner(ws, row, "REQUEST", FILL_REQUEST, ncols); row += 1
    style_column_header(ws, row); row += 1

    params = op.get("parameters", [])
    header_params = [p for p in params if p.get("in") == "header"]
    body_params = [p for p in params if p.get("in") == "body"]
    query_params = [p for p in params if p.get("in") in ("query", "path", "formData")]

    style_subgroup_row(ws, row, "header (헤더)", ncols); row += 1
    consumes = op.get("consumes") or ["application/json"]
    write_field_row(ws, row, 0, "Content-Type", "", "string", "", "Y", consumes[0]); row += 1
    for p in header_params:
        write_field_row(ws, row, 0, p["name"], "", "string", "", "Y" if p.get("required") else "N", p.get("description", "")); row += 1

    if query_params:
        style_subgroup_row(ws, row, "query / path (파라미터)", ncols); row += 1
        for p in query_params:
            write_field_row(ws, row, 0, p["name"], "", type_label(p, definitions), p.get("maxLength", ""),
                             "Y" if p.get("required") else "N", p.get("description", "")); row += 1

    style_subgroup_row(ws, row, "body (바디)", ncols); row += 1
    for p in body_params:
        for level, name, definition, dtype, size, required, desc in flatten_schema(p.get("schema", {}), definitions):
            write_field_row(ws, row, level, name, definition, dtype, size, required, desc); row += 1

    apply_outer_border(ws, request_start, row - 1, C(1), C(ncols))
    row += 1  # blank line

    # RESPONSE
    response_start = row
    style_section_banner(ws, row, "RESPONSE", FILL_RESPONSE, ncols); row += 1
    style_column_header(ws, row); row += 1

    style_subgroup_row(ws, row, "status (상태)", ncols); row += 1
    write_field_row(ws, row, 0, "HTTP/1.1 200 OK", "정상처리시 200 상태", "", "", "Y", "서버측 응답 Http 결과 상태"); row += 1

    style_subgroup_row(ws, row, "body (바디)", ncols); row += 1
    responses = op.get("responses", {})
    ok = responses.get("200", {})
    schema = ok.get("schema")
    if schema:
        for level, name, definition, dtype, size, required, desc in flatten_schema(schema, definitions):
            write_field_row(ws, row, level, name, definition, dtype, size, required, desc); row += 1

    apply_outer_border(ws, response_start, row - 1, C(1), C(ncols))

    ws.freeze_panes = "B9"
    return ws


def build_toc_sheet(wb: Workbook, tag: str, entries: list, used_names: set):
    ws = wb.create_sheet(sanitize_sheet_name(tag, used_names))
    ncols = 6
    ws.column_dimensions["A"].width = 3
    ws.row_dimensions[1].height = 6
    headers = ["No", "API명", "Method", "URL", "설명", "이동"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=C(c), value=h)
        cell.font = BOLD
        cell.fill = FILL_HEADER_ROW
        cell.border = BORDER
    widths = [6, 30, 10, 45, 40, 10]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(C(i))].width = w

    for i, (no, name, method, path, desc, sheet_name) in enumerate(entries, start=3):
        ws.cell(row=i, column=C(1), value=no).border = BORDER
        ws.cell(row=i, column=C(2), value=name).border = BORDER
        ws.cell(row=i, column=C(3), value=method.upper()).border = BORDER
        ws.cell(row=i, column=C(4), value=path).border = BORDER
        ws.cell(row=i, column=C(5), value=desc).border = BORDER
        link_cell = ws.cell(row=i, column=C(6), value="바로가기")
        link_cell.hyperlink = f"#'{sheet_name}'!A1"
        link_cell.font = Font(color="0000FF", underline="single")
        link_cell.border = BORDER
    apply_outer_border(ws, 2, len(entries) + 2, C(1), C(ncols))
    ws.freeze_panes = "B3"
    return ws


def build_history_sheet(wb: Workbook, today: str):
    ws = wb.create_sheet("업데이트 이력")
    ncols = 4
    ws.column_dimensions["A"].width = 3
    ws.row_dimensions[1].height = 6
    headers = ["버전", "날짜", "작성자", "변경내용"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=C(c), value=h)
        cell.font = BOLD
        cell.fill = FILL_HEADER_ROW
        cell.border = BORDER
    data_row = ["1.0", today, "", "최초 생성 (Swagger 자동 변환)"]
    for c, v in enumerate(data_row, start=1):
        ws.cell(row=3, column=C(c), value=v).border = BORDER
    widths = [10, 14, 14, 50]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(C(i))].width = w
    apply_outer_border(ws, 2, 3, C(1), C(ncols))
    return ws


def build_tag_workbook(tag: str, tag_operations: list, definitions: dict, today: str) -> Workbook:
    """태그(컨트롤러) 하나에 대한 워크북(업데이트이력 + 목차 + 상세 시트들)을 만든다."""
    used_names = set()
    named_ops = []
    toc_entries = []
    for no, (path, method, op) in enumerate(tag_operations, start=1):
        title = op.get("summary") or path.rstrip("/").split("/")[-1]
        sheet_name = sanitize_sheet_name(f"{no}. {title}", used_names)
        named_ops.append((no, path, method, op, sheet_name))
        toc_entries.append((no, title, method, path, op.get("description", ""), sheet_name))

    wb = Workbook()
    wb.remove(wb.active)
    build_history_sheet(wb, today)

    toc_used_names = set(used_names)
    build_toc_sheet(wb, tag, toc_entries, toc_used_names)

    for no, path, method, op, sheet_name in named_ops:
        build_detail_sheet(wb, sheet_name, tag, path, method, op, definitions)

    return wb


def group_operations_by_tag(spec: dict, tag_filter=None, limit=None) -> dict:
    paths = spec.get("paths", {})
    operations = []
    for path, methods in paths.items():
        for method, op in methods.items():
            tag = (op.get("tags") or ["기타"])[0]
            if tag_filter and tag not in tag_filter:
                continue
            operations.append((tag, path, method, op))

    operations.sort(key=lambda x: (x[0], x[1]))
    if limit:
        operations = operations[:limit]

    by_tag = {}
    for tag, path, method, op in operations:
        by_tag.setdefault(tag, []).append((path, method, op))
    return by_tag


def main():
    parser = argparse.ArgumentParser(description="Swagger 2.0 JSON -> 태그(컨트롤러)별 API 정의서 엑셀 변환")
    parser.add_argument("--url", required=True, help="Swagger JSON 엔드포인트 URL")
    parser.add_argument("--output-dir", required=True, help="출력 xlsx 파일들을 저장할 디렉토리")
    parser.add_argument("--prefix", required=True, help="파일명 접두사 (예: FMRK_WAVEN) -> {번호}.{접두사}_{태그}_api정의서.xlsx")
    parser.add_argument("--tags", nargs="*", help="특정 tag만 변환 (샘플 생성용)")
    parser.add_argument("--limit", type=int, help="처음 N개 API만 변환 (샘플 생성용)")
    args = parser.parse_args()

    import datetime
    import os

    today = datetime.date.today().isoformat()
    os.makedirs(args.output_dir, exist_ok=True)

    spec = fetch_swagger(args.url)
    definitions = spec.get("definitions", {})
    by_tag = group_operations_by_tag(spec, tag_filter=set(args.tags) if args.tags else None, limit=args.limit)

    for no, tag in enumerate(sorted(by_tag), start=1):
        wb = build_tag_workbook(tag, by_tag[tag], definitions, today)
        tag_label = re.sub(r"Api$", "", tag) or tag
        filename = f"{no}.{args.prefix}_{tag_label}_api정의서.xlsx"
        out_path = os.path.join(args.output_dir, filename)
        wb.save(out_path)
        print(f"저장 완료: {out_path}")


if __name__ == "__main__":
    main()
